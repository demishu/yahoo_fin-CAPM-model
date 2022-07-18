# -*- coding: utf-8 -*-
"""
Created on Sat Jul 16 18:17:25 2022

@author: demishu
"""

import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter
import openpyxl
from yahoo_fin import stock_info as si
import datetime
import os
from scipy.stats import linregress as ols

ymd = datetime.datetime.now().strftime('%Y-%m-%d')

def set_interval(first_day = None, last_day = None, interval = 365, fmt="%Y/%m/%d"):
    """set_interval(第一天,最后一天，间隔，格式)-->     (第一天，最后一天)     \n
        函数可以只输入(第一天)或者(间隔)-->           （第一天，今天）          \n
        默认返回                                     （去年的今天，今天）      \n
        fmt参数用于将输入的日期格式转换成datetime对象，然后统一输出为 "%Y/%m/%d"格式
    """
    if last_day is None:
        last_day = datetime.datetime.now()
    else:
        try:
            last_day = datetime.datetime.strptime(last_day, fmt)
        except:
            raise("检查输入的最后一天")
    if first_day:
        try:
            first_day = datetime.datetime.strptime(first_day,fmt).strftime("%Y/%m/%d")
            last_day = last_day.strftime("%Y/%m/%d")
            return first_day, last_day
        except:
            raise('检查输入的第一天。')
    elif isinstance(interval,int):
        first_day = last_day - datetime.timedelta(interval)
        first_day = first_day.strftime("%Y/%m/%d")
        last_day = last_day.strftime("%Y/%m/%d")
        return first_day,last_day

class Meta_output_class:
    def _set_writer(self,output_path):
        self._writer = pd.ExcelWriter(output_path, mode = 'a', engine = 'openpyxl',
                                if_sheet_exists='replace')
    def _create_blank_sheet(self):
        if not os.path.exists(self._output_path):
            print("未检测到输出文件，正在创建")
            if not os.path.exists(os.path.dirname(self._output_path)):
                os.makedirs(os.path.dirname(self._output_path))
            df = pd.DataFrame()
            df.to_excel(self._output_path, sheet_name='一串乱码')
            print("创建成功\n")
            self._blank_sheet = True
        else:
            self._blank_sheet = False

    def _set_columns_width(self, df, sheet):
        print("正在设置列宽")
        df = pd.DataFrame(df)
        df.loc['column',:] = df.columns
        widths = df.astype(str).applymap(lambda x: len(str(x).encode('utf-8'))).agg(max).values
        worksheet = self._writer.sheets[sheet]
        for i, width in enumerate(widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width
        print("设置完成")
        return worksheet
    def _clean_blank_sheet(self):
        if self._blank_sheet is True:
            print("检测到空白sheet")
            wb=openpyxl.load_workbook(self._output_path)
            del wb["一串乱码"]
            print("删除成功")
            wb.save(self._output_path)
            wb.close()
            self._blank_sheet = False

class CAPM(Meta_output_class):
    def __init__(self,tickers_list,start_day, end_day, output_data = False):
        self._output_path = f'./CAPM模型 {ymd}.xlsx'
        self._output_data = output_data
        self._tickers_list = tickers_list
        self._start_day = start_day
        self._end_day = end_day
        
        self._main()
        
    "__init__里设置类属性，_main跑实际的类方法"
    def _main(self):
        self._get_data()
        self._ols_processing()
    
    "类装饰器"
    def _to_excel(func):
        def new_func(self, *args, **kwargs):
            df = func(self, *args, **kwargs)
            if self._output_data:
                print(f"开始输出{self._output_path}")
                self._create_blank_sheet()
                self._set_writer(self._output_path)
                print(f"正在保存 {self._sheet_name}工作表")
                df.to_excel(self._writer, self._sheet_name)
                self._set_columns_width(df,self._sheet_name)
                self._writer.save()
                self._writer.close()
                del self._sheet_name     
                self._clean_blank_sheet()
        return new_func
    '函数'
    def _OLS(self, y, x):
        "scipy的最小二乘法OLS"
        slope, intercept, r_value, p_value, std_err = ols(x, y)
        return intercept, slope, r_value**2, p_value
    '属性'
    @property
    def data(self):
        return self._data
    @property
    def stats(self):
        return self._stats
    
    "类方法 -- 开头"
    @_to_excel
    def _get_data(self):
        self._sheet_name = '收盘数据'
        """调用yahoo_fin来爬取股票、指数信息"""
        tickers_dict={}
        data_df =pd.DataFrame()
        
        for i, ticker in enumerate(self._tickers_list, start = 1):
            print(f"进度条：{i}/{len(self._tickers_list)}\n开始读取{ticker}")
            df = si.get_data(ticker, self._start_day, self._end_day, index_as_date= True)
            tickers_dict[ticker] = df
            
            if ticker == '^IXIC':
                data_df["纳指"] = df ["adjclose"]
            else:
                print(f"正在处理{ticker}\n")
                data_df[ticker] = df.loc[:,"adjclose"]
        "股票交易数据，以备不时之需（暂时用不上）"
        self._tickers_dict = tickers_dict
        "股票收盘信息的df，后续用得上"
        self._data = data_df
        return data_df

    @_to_excel
    def _ols_processing(self):
        self._sheet_name = 'CAPM模型'
        '预处理'
        df = pd.DataFrame(self._data)
        df = df.pct_change()
        df = df.dropna()
        '计算OLS'
        stats = df.apply (lambda x: self._OLS(df["纳指"], x) )
        stats.index = ["alpha/intercept",'beta/slope','Rsq','P-value']
        stats.loc['mean',:] = df.mean()
        
        """
        ri      =   股票i的平均收益       (mean行)
        betai   =   股票i的beta          (beta/slope行)
        rM      =   市场的收益           （stats.loc['mean','纳指'])
        rf      =   无风险收益
        由于ri = rf+betai(rM-rf)
        可推导出
        ri - betai * rM = rf*(1-betai)
        设等式左边为y， 右半边则是rf*x
        需要再OLS一次，获得rf的信息
        """
        rf_data = list(self._OLS(stats["mean"]-stats['beta/slope']*stats.loc['纳指','mean'],
                                    1-stats['beta/slope']))
        '需要新增一行nan，因为rf这一列没有mean。 如果不添加nan， pandas就会因为\
        stats有5行，而rf_data只有4行，而insert失败'
        rf_data.append(np.nan)
        stats.insert(0,'rf',rf_data)
        stats = stats.T
        self._stats = stats
        return stats
    '类方法 -- 结尾'
    
if __name__=='__main__':
    first_day, last_day = set_interval('2022/1/1')
    ticker_list = list()                #列表必须包含纳指（"^IXIC")。 假设我想爬苹果和微软的股票，那么我就要写['^IXIC','AAPL','MSFT']。 股票必须以股票代码的形式写入列表。
    model = CAPM(ticker_list,first_day, last_day, output_data = True)
